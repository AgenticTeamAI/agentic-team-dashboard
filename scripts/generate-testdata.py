#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Genereert de fictieve testbundel voor het Agentic Team Dashboard, in alle drie
de formaten die het dashboard moet kunnen lezen: Excel-werkboek, data/*.json
en een Notion-export-map. Fictieve klant: "GroenBuro" (kantoorbeplanting en
werkplekwelzijn voor MKB) - geen enkele overeenkomst met een echte klant.

Veldnamen komen 1-op-1 uit schema/schema.generated.js (zelf afgeleid uit
core/agents.json -> datadomeinen), nooit hier opnieuw verzonnen.

Bewuste gaten, verspreid over de drie bundels (zie README testdata/ voor een
overzicht per bundel):
 - een leeg domein (0 rijen, wel het bestand/tabblad/sheet zelf)
 - een verouderd domein of hele bundel (bestandsdatum ouder dan de
   veroudering-drempel van 30 dagen)
 - een ontbrekend domein in een verder complete bundel
 - een agent die in de Acties/Lessen-sporen geen enkele vermelding heeft
 - een bedrijfscontext-bestand dat volledig ontbreekt, dat stale is, en dat
   compleet is - drie verschillende zone 2-situaties over de drie bundels

Alle namen, bedragen en gebeurtenissen zijn verzonnen.
"""
import json
import os
import time
from datetime import date, datetime, timedelta
from pathlib import Path

try:
    from openpyxl import Workbook
except ImportError:
    raise SystemExit("openpyxl ontbreekt. Installeer met: pip install openpyxl")

ROOT = Path(__file__).parent.parent
TESTDATA = ROOT / "testdata"
TODAY = date(2026, 8, 10)  # vaste "vandaag" zodat de bundel reproduceerbaar is


def d(offset_days):
    return (TODAY + timedelta(days=offset_days)).isoformat()


# ─────────────────────────────────────────────────────────────────────────
# Canonieke datasets per domein. Sleutels zijn EXACT de veldnamen uit de
# registry (schema/schema.generated.js), zodat elke route (Excel/JSON/Notion)
# dezelfde kolomnamen gebruikt - conform het ontwerp ("dezelfde velden
# ongeacht bron").
# ─────────────────────────────────────────────────────────────────────────

ORGANISATIES = [
    {
        "Naam": "Kantoorpark De Linde", "Fase": "Actieve klant", "Segment": "Kantoren 50-250 medewerkers",
        "Omvang FTE": "100-250", "Lead Bron": "Netwerk", "Signaal": "Nieuw pand betrokken, wil plantenwand in ontvangstruimte",
        "Signaal Datum": d(-120), "Website": "https://kantoorparkdelinde-fictief.nl", "LinkedIn": "",
        "Vestigingsplaats": "Amersfoort", "Eigenaar": "Nina de Groot", "KvK-nummer": "61234501", "Notities": "Loopt sinds 2025, tevreden klant.",
    },
    {
        "Naam": "Buro Helder Advocaten", "Fase": "Lead", "Segment": "Kantoren <50 medewerkers",
        "Omvang FTE": "<50", "Lead Bron": "LinkedIn", "Signaal": "Directiesecretaresse vroeg naar abonnementsvormen",
        "Signaal Datum": d(-14), "Website": "https://burohelder-fictief.nl", "LinkedIn": "https://linkedin.com/company/burohelder-fictief",
        "Vestigingsplaats": "Utrecht", "Eigenaar": "Nina de Groot", "KvK-nummer": "61234502", "Notities": "Eerste gesprek gepland.",
    },
    {
        "Naam": "CoWorkPlek Zuid", "Fase": "Prospect", "Segment": "Coworking spaces",
        "Omvang FTE": "50-100", "Lead Bron": "Event", "Signaal": "Ontmoet op Werkplek Vakbeurs, zoekt frisse uitstraling",
        "Signaal Datum": d(-30), "Website": "", "LinkedIn": "",
        "Vestigingsplaats": "Rotterdam", "Eigenaar": "Sam Kuiper", "KvK-nummer": "", "Notities": "Nog geen vervolgcontact geweest.",
    },
    {
        "Naam": "Zorggroep Vredehof", "Fase": "Actieve klant", "Segment": "Zorginstellingen",
        "Omvang FTE": "250-500", "Lead Bron": "Referral", "Signaal": "Wil groen in 3 wachtruimtes voor rust bij bezoekers",
        "Signaal Datum": d(-200), "Website": "https://vredehof-fictief.nl", "LinkedIn": "",
        "Vestigingsplaats": "Zwolle", "Eigenaar": "Nina de Groot", "KvK-nummer": "61234504", "Notities": "Contract loopt door tot 2027.",
    },
    {
        "Naam": "StartupHub Noord", "Fase": "Voormalig klant", "Segment": "Coworking spaces",
        "Omvang FTE": "<50", "Lead Bron": "Inbound", "Signaal": "Abonnement opgezegd na verhuizing",
        "Signaal Datum": d(-260), "Website": "", "LinkedIn": "",
        "Vestigingsplaats": "Groningen", "Eigenaar": "Sam Kuiper", "KvK-nummer": "", "Notities": "Mogelijk terug bij nieuwe locatie.",
    },
]

CONTACTPERSONEN = [
    {"Naam": "Renate Bosman", "Functie": "Office Manager", "E-mail": "r.bosman@kantoorparkdelinde-fictief.nl", "LinkedIn URL": "",
     "Warmte": "Heet", "Outreach Status": "In gesprek", "Rol in besluitvorming": "Champion", "Laatste Contact": d(-5), "Notities": "Erg enthousiast, wil uitbreiden."},
    {"Naam": "Farid Al-Amin", "Functie": "Directiesecretaresse", "E-mail": "f.alamin@burohelder-fictief.nl", "LinkedIn URL": "",
     "Warmte": "Warm", "Outreach Status": "Gereageerd", "Rol in besluitvorming": "Gatekeeper", "Laatste Contact": d(-8), "Notities": "Wil offerte voor 2 verdiepingen."},
    {"Naam": "Els Timmerman", "Functie": "Locatiemanager", "E-mail": "", "LinkedIn URL": "",
     "Warmte": "Koud", "Outreach Status": "Niet benaderd", "Rol in besluitvorming": "Decision Maker", "Laatste Contact": "", "Notities": "Nog niet benaderd sinds de beurs."},
    {"Naam": "Petra van Es", "Functie": "Facilitair coördinator", "E-mail": "p.vanes@vredehof-fictief.nl", "LinkedIn URL": "",
     "Warmte": "Warm", "Outreach Status": "In gesprek", "Rol in besluitvorming": "Influencer", "Laatste Contact": d(-20), "Notities": "Wacht op akkoord van bestuur."},
    {"Naam": "Joost Rademaker", "Functie": "Ex-locatiemanager", "E-mail": "", "LinkedIn URL": "",
     "Warmte": "Koud", "Outreach Status": "Niet reageren", "Rol in besluitvorming": "Decision Maker", "Laatste Contact": d(-260), "Notities": "Gaf aan geen interesse meer te hebben."},
]

INTERACTIES = [
    {"Onderwerp": "Kennismaking plantenwand", "Type": "Meeting", "Richting": "Uitgaand", "Datum": d(-5),
     "Samenvatting": "Renate wil offerte voor ontvangstruimte + 2 vergaderzalen.", "Sentiment": "Positief",
     "Volgende Actie": "Offerte opstellen", "Actie Deadline": d(2)},
    {"Onderwerp": "Vervolgvraag abonnementsvormen", "Type": "E-mail", "Richting": "Inkomend", "Datum": d(-8),
     "Samenvatting": "Farid vraagt naar prijsverschil vast vs. per kwartaal.", "Sentiment": "Neutraal",
     "Volgende Actie": "Prijsoverzicht sturen", "Actie Deadline": d(-3)},
    {"Onderwerp": "Beursgesprek Werkplek Vakbeurs", "Type": "Event", "Richting": "Inkomend", "Datum": d(-30),
     "Samenvatting": "Kort gesprek bij de stand, visitekaartje uitgewisseld.", "Sentiment": "Neutraal",
     "Volgende Actie": "Opvolgmail sturen", "Actie Deadline": d(-25)},
    {"Onderwerp": "Update wachtruimtes", "Type": "Telefoongesprek", "Richting": "Uitgaand", "Datum": d(-20),
     "Samenvatting": "Petra meldt dat bestuur nog moet beslissen over budget.", "Sentiment": "Neutraal",
     "Volgende Actie": "Over 2 weken terugbellen", "Actie Deadline": d(-6)},
]

SALES_FUNNEL = [
    {"Deal Naam": "De Linde - uitbreiding 2 vergaderzalen", "Fase": "Oplossingsrichting (70%)", "Opvolg Status": "Op schema",
     "Verwachte Omzet": 4200, "Probability": 0.70, "Verwachte Sluitdatum": d(20), "Volgende Actie": "Offerte versturen",
     "Volgende Actie Deadline": d(2), "Laatste Contact": d(-5), "Lead Bron": "Netwerk", "Eigenaar": "Nina de Groot", "Notities": "Warme klant, uitbreiding op bestaand contract."},
    {"Deal Naam": "Buro Helder - startpakket 2 verdiepingen", "Fase": "Probleemdefinitie (30%)", "Opvolg Status": "Actie Vereist",
     "Verwachte Omzet": 3100, "Probability": 0.30, "Verwachte Sluitdatum": d(45), "Volgende Actie": "Prijsoverzicht sturen",
     "Volgende Actie Deadline": d(-3), "Laatste Contact": d(-8), "Lead Bron": "LinkedIn", "Eigenaar": "Nina de Groot", "Notities": "Deadline al verstreken, actie blijft liggen."},
    {"Deal Naam": "CoWorkPlek Zuid - oriëntatie", "Fase": "Targeting (10%)", "Opvolg Status": "Actie Vereist",
     "Verwachte Omzet": 1800, "Probability": 0.10, "Verwachte Sluitdatum": d(90), "Volgende Actie": "Opvolgmail sturen",
     "Volgende Actie Deadline": d(-25), "Laatste Contact": d(-30), "Lead Bron": "Event", "Eigenaar": "Sam Kuiper", "Notities": "Al een maand stil, geen reactie op eerste mail."},
    {"Deal Naam": "Vredehof - 3 wachtruimtes", "Fase": "Besluitvorming (85%)", "Opvolg Status": "In afwachting",
     "Verwachte Omzet": 6800, "Probability": 0.85, "Verwachte Sluitdatum": d(15), "Volgende Actie": "Terugbellen na bestuursbesluit",
     "Volgende Actie Deadline": d(-6), "Laatste Contact": d(-20), "Lead Bron": "Referral", "Eigenaar": "Nina de Groot", "Notities": "Wacht op bestuur, buiten onze controle."},
    {"Deal Naam": "StartupHub Noord - heraansluiting", "Fase": "Discovery (20%)", "Opvolg Status": "Verloren",
     "Verwachte Omzet": 900, "Probability": 0.0, "Verwachte Sluitdatum": d(-40), "Volgende Actie": "",
     "Volgende Actie Deadline": "", "Laatste Contact": d(-90), "Lead Bron": "Inbound", "Eigenaar": "Sam Kuiper", "Notities": "Definitief geen budget dit jaar."},
]

ACTIES = [
    {"Actie": "Offerte De Linde opstellen", "Eigenaar": "Nina de Groot", "Agent": "outreach-specialist", "Deadline": d(2), "Status": "Bezig"},
    {"Actie": "Prijsoverzicht Buro Helder sturen", "Eigenaar": "Nina de Groot", "Agent": "outreach-specialist", "Deadline": d(-3), "Status": "Open"},
    {"Actie": "QC-bevinding: outreachmail refereert verkeerd segment", "Eigenaar": "Nina de Groot", "Agent": "quality-control", "Deadline": d(-1), "Status": "Open"},
    {"Actie": "QC-bevinding: dubbele deal-entry gesignaleerd", "Eigenaar": "Sam Kuiper", "Agent": "quality-control", "Deadline": d(1), "Status": "Open"},
    {"Actie": "Maandrapportage omzet voorbereiden", "Eigenaar": "Sam Kuiper", "Agent": "controller", "Deadline": d(-10), "Status": "Klaar"},
    {"Actie": "Contentkalender augustus vullen", "Eigenaar": "Nina de Groot", "Agent": "content-strateeg", "Deadline": d(5), "Status": "Bezig"},
    {"Actie": "Check-in Vredehof inplannen", "Eigenaar": "Nina de Groot", "Agent": "customer-success-manager", "Deadline": d(3), "Status": "Open"},
    {"Actie": "Vervolgmail CoWorkPlek Zuid", "Eigenaar": "Sam Kuiper", "Agent": "researcher", "Deadline": d(-25), "Status": "Open"},
]

LESSEN_INZICHTEN = [
    {"Les": "Outreachmails naar zorginstellingen werken beter met concreet rustvoorbeeld dan met prijs vooraan.",
     "Agent": "outreach-specialist", "Categorie": "Outreach & Acquisitie", "Datum": d(-40), "Actie": "Sjabloon aangepast voor zorgsegment.",
     "Status": "Afgerond", "Impact": "Medium", "Verwerkt in prompt?": True},
    {"Les": "Klanten in Coworking-segment reageren traag op e-mail, wel snel op telefoon.",
     "Agent": "researcher", "Categorie": "Markt & Prospects", "Datum": d(-33), "Actie": "Belscript toevoegen aan playbook.",
     "Status": "Open", "Impact": "Medium", "Verwerkt in prompt?": False},
    {"Les": "Facturatie liep vertraging op doordat contractwaarde niet was ingevuld bij projectstart.",
     "Agent": "controller", "Categorie": "Operations & Finance", "Datum": d(-60), "Actie": "Verplicht veld bij intake benoemen.",
     "Status": "In uitvoering", "Impact": "Hoog", "Verwerkt in prompt?": False},
    {"Les": "QC signaleerde tweemaal een outreachmail met verkeerd ingevuld segment - bronveld stond niet aan het begin van de playbook-fase.",
     "Agent": "quality-control", "Categorie": "Kwaliteit & Review", "Datum": d(-2), "Actie": "Werkwijze vooraan in outreach-playbook zetten.",
     "Status": "Open", "Impact": "Hoog", "Verwerkt in prompt?": False},
    {"Les": "Klantsuccescheck bij Vredehof te laat ingepland na signaal van dalende tevredenheid.",
     "Agent": "customer-success-manager", "Categorie": "Klantsucces & Retentie", "Datum": d(-15), "Actie": "Escalatiedrempel health-status verlagen.",
     "Status": "Open", "Impact": "Hoog", "Verwerkt in prompt?": False},
    {"Les": "Contentkalender liep leeg in juli door vakantie - geen vervangend proces.",
     "Agent": "content-strateeg", "Categorie": "Content & Zichtbaarheid", "Datum": d(-25), "Actie": "Back-up eigenaar aanwijzen voor kalender.",
     "Status": "Vervallen", "Impact": "Laag", "Verwerkt in prompt?": False},
]

DAGVERSLAGEN = [
    {"Naam": "Dagverslag 4 augustus", "Type": "Dagverslag", "Dag": d(-6), "Status": "Afgerond", "Persoon": "Nina de Groot"},
    {"Naam": "Weekoverzicht week 31", "Type": "Weekoverzicht", "Dag": d(-10), "Status": "Afgerond", "Persoon": "Nina de Groot"},
    {"Naam": "Dagverslag 8 augustus", "Type": "Dagverslag", "Dag": d(-2), "Status": "Afgerond", "Persoon": "Sam Kuiper"},
    {"Naam": "Dagverslag 10 augustus", "Type": "Dagverslag", "Dag": d(0), "Status": "In uitvoering", "Persoon": "Nina de Groot"},
]

PRODUCTBACKLOG = [
    {"Item": "Automatische seizoensattendering (plantwissel voorjaar/winter)", "Herkomst": "Klantsucces-signalen", "Eigenaar": "Product Designer",
     "Status": "Doing", "Prioriteit": "Hoog", "Besluit": "Opgenomen in Q3-planning."},
    {"Item": "Losse consultafspraak zonder abonnement", "Herkomst": "Lead-gesprekken", "Eigenaar": "Product Designer",
     "Status": "Ter validatie", "Prioriteit": "Midden", "Besluit": ""},
    {"Item": "Kwartaalabonnement voor coworking-segment", "Herkomst": "Sales-signalen", "Eigenaar": "Delivery Architect",
     "Status": "To do", "Prioriteit": "Laag", "Besluit": ""},
    {"Item": "Duurzaamheidsrapportage per klant", "Herkomst": "Klantvraag Vredehof", "Eigenaar": "Product Designer",
     "Status": "Done", "Prioriteit": "Hoog", "Besluit": "Gebouwd en uitgerold in juni."},
]

KLANTSUCCES = [
    {"Klantnaam": "Kantoorpark De Linde", "Fase": "Actief", "Health": "Groen", "Verlengdatum": d(140), "Laatste check": d(-5),
     "Signalen": "Uitbreiding aangevraagd, positief.", "Eigenaar": "Nina de Groot"},
    {"Klantnaam": "Zorggroep Vredehof", "Fase": "Risico", "Health": "Rood", "Verlengdatum": d(60), "Laatste check": d(-20),
     "Signalen": "Dalende tevredenheid, escalatie te laat opgepakt.", "Eigenaar": "Nina de Groot"},
    {"Klantnaam": "StartupHub Noord", "Fase": "Verloren", "Health": "Rood", "Verlengdatum": "", "Laatste check": d(-260),
     "Signalen": "Opgezegd na verhuizing.", "Eigenaar": "Sam Kuiper"},
    {"Klantnaam": "Facilitair Collectief Oost", "Fase": "Verlenging", "Health": "Oranje", "Verlengdatum": d(25), "Laatste check": d(-35),
     "Signalen": "Nog geen reactie op verlengingsvoorstel.", "Eigenaar": "Sam Kuiper"},
]

PROJECTEN = [
    {"Projectnaam": "De Linde - inrichting ontvangstruimte", "Status": "Actief", "Startdatum": d(-40), "Einddatum": d(20),
     "Contractwaarde": 4200, "Gefactureerd": 2100, "Eigenaar": "Nina de Groot", "Gedeelde Pagina": "", "Notities": "Fase 1 opgeleverd."},
    {"Projectnaam": "Vredehof - 3 wachtruimtes", "Status": "Opstarten", "Startdatum": d(10), "Einddatum": d(70),
     "Contractwaarde": 6800, "Gefactureerd": 0, "Eigenaar": "Nina de Groot", "Gedeelde Pagina": "", "Notities": "Wacht op bestuursbesluit."},
    {"Projectnaam": "StartupHub Noord - onderhoud 2025", "Status": "Afgerond", "Startdatum": d(-400), "Einddatum": d(-260),
     "Contractwaarde": 1200, "Gefactureerd": 1200, "Eigenaar": "Sam Kuiper", "Gedeelde Pagina": "", "Notities": "Contract niet verlengd."},
]

OFFERTES = [
    {"Offertenummer": "OFF-2026-014", "Status": "Verstuurd", "Bedrag excl. BTW": 4200, "BTW bedrag": 882, "Geldig tot": d(14),
     "Verstuurd op": d(-1), "Getekend op": "", "Offerte URL": "", "Notities": "De Linde - uitbreiding."},
    {"Offertenummer": "OFF-2026-011", "Status": "Geaccepteerd", "Bedrag excl. BTW": 6800, "BTW bedrag": 1428, "Geldig tot": d(-5),
     "Verstuurd op": d(-25), "Getekend op": d(-18), "Offerte URL": "", "Notities": "Vredehof, wacht nog op startdatum."},
    {"Offertenummer": "OFF-2026-009", "Status": "Verlopen", "Bedrag excl. BTW": 900, "BTW bedrag": 189, "Geldig tot": d(-30),
     "Verstuurd op": d(-70), "Getekend op": "", "Offerte URL": "", "Notities": "StartupHub Noord, geen reactie meer gekomen."},
]

CONTENT_KALENDER = [
    {"Titel": "5 planten die luchtvochtigheid in kantoren verbeteren", "Status": "Gepubliceerd", "Type": "Artikel",
     "Segment": "Kantoren 50-250 medewerkers", "Publicatiedatum": d(-12), "Kanaal": ["Website", "LinkedIn"],
     "Hook A": "Uw airco droogt de lucht meer uit dan u denkt.", "Hook B": "", "Concept Tekst": "", "Campagne": "Zomer 2026", "Eigenaar": "Nina de Groot"},
    {"Titel": "Case: Zorggroep Vredehof over rust in de wachtruimte", "Status": "Klaar voor review", "Type": "Case study",
     "Segment": "Zorginstellingen", "Publicatiedatum": d(6), "Kanaal": ["Website"],
     "Hook A": "", "Hook B": "", "Concept Tekst": "Concept in review bij klant.", "Campagne": "", "Eigenaar": "Nina de Groot"},
    {"Titel": "LinkedIn-post: onderhoudstip augustus", "Status": "Gepland", "Type": "LinkedIn post",
     "Segment": "Coworking spaces", "Publicatiedatum": d(3), "Kanaal": ["LinkedIn"],
     "Hook A": "Uw plant overleeft de vakantie niet zonder dit.", "Hook B": "", "Concept Tekst": "", "Campagne": "", "Eigenaar": "Nina de Groot"},
    {"Titel": "Idee: kort filmpje plantverzorging voor de balie", "Status": "Idee", "Type": "Video",
     "Segment": "Kantoren <50 medewerkers", "Publicatiedatum": "", "Kanaal": [],
     "Hook A": "", "Hook B": "", "Concept Tekst": "", "Campagne": "", "Eigenaar": "Sam Kuiper"},
    {"Titel": "E-mail: seizoenswissel voorjaar", "Status": "Archief", "Type": "E-mail template",
     "Segment": "Kantoren 50-250 medewerkers", "Publicatiedatum": d(-150), "Kanaal": ["E-mail"],
     "Hook A": "", "Hook B": "", "Concept Tekst": "", "Campagne": "Voorjaar 2026", "Eigenaar": "Nina de Groot"},
]

DELIVERY_RUGZAK = [
    {"Naam": "Intakeformulier plantenwand", "Type": "Template / Tool", "Status": "Gereed",
     "Beschrijving": "Vragenlijst voor lichtinval, ruimte en onderhoudsvoorkeur.", "Wanneer inzetten": "Bij elke nieuwe klant",
     "Benodigdheden": "", "Segment": [], "Bron": ""},
    {"Naam": "Seizoenschecklist", "Type": "Werkvorm / Oefening", "Status": "In ontwikkeling",
     "Beschrijving": "Checklist voor plantwissel per seizoen.", "Wanneer inzetten": "Bij seizoensovergang",
     "Benodigdheden": "", "Segment": [], "Bron": ""},
    {"Naam": "Rust-in-de-wachtruimte model", "Type": "Model / Framework", "Status": "💡 Idee",
     "Beschrijving": "Raamwerk voor inrichting van wachtruimtes in de zorg.", "Wanneer inzetten": "Bij zorginstellingen",
     "Benodigdheden": "", "Segment": [], "Bron": ""},
]

TIJDREGISTRATIE = [
    {"Beschrijving": "Inrichting ontvangstruimte De Linde", "Datum": d(-6), "Uren": 4.5, "Type": "Billable", "Uurtarief": 65, "Gefactureerd": True, "Persoon": "Nina de Groot"},
    {"Beschrijving": "Voorbereiding offerte Vredehof", "Datum": d(-20), "Uren": 2.0, "Type": "Billable", "Uurtarief": 65, "Gefactureerd": False, "Persoon": "Nina de Groot"},
    {"Beschrijving": "Interne planning augustus", "Datum": d(-3), "Uren": 3.0, "Type": "Intern", "Uurtarief": 0, "Gefactureerd": False, "Persoon": "Sam Kuiper"},
    {"Beschrijving": "Contentkalender vullen", "Datum": d(-1), "Uren": 1.5, "Type": "Marketing", "Uurtarief": 0, "Gefactureerd": False, "Persoon": "Nina de Groot"},
]

PRODUCT_CATALOGUS = [
    {"Productnaam": "Basis Plantenabonnement", "Status": "Actief", "Type": "Abonnement", "Prijs": 89, "Prijsmodel": "Per maand",
     "Plek in ladder": "Activatie", "Segment": ["Kantoren <50 medewerkers"], "Beschrijving": "Maandelijks onderhoud van 5 planten.",
     "USP": "Geen zorgen, wij verzorgen.", "Notities": ""},
    {"Productnaam": "Plantenwand op maat", "Status": "Actief", "Type": "Advies", "Prijs": 0, "Prijsmodel": "Op maat",
     "Plek in ladder": "Verdieping", "Segment": ["Kantoren 50-250 medewerkers", "Zorginstellingen"], "Beschrijving": "Ontwerp en plaatsing van een groene wand.",
     "USP": "Statement piece voor de ontvangstruimte.", "Notities": ""},
    {"Productnaam": "Seizoensscan", "Status": "In creatie", "Type": "Assessment", "Prijs": 0, "Prijsmodel": "Vast",
     "Plek in ladder": "Diagnose", "Segment": ["Coworking spaces"], "Beschrijving": "Eenmalige scan van lichtinval en klimaat.",
     "USP": "", "Notities": "Nog in ontwikkeling."},
]

BEDRIJFSCONTEXT_COMPLEET = {
    "_schema_opmerking": "VOORLOPIG - dit domein staat nog niet in core/agents.json -> datadomeinen. "
                          "Formaat hangt af van de S17/f13-uitwerking (zie BESLUIT-s17). Dit dashboard "
                          "toont deze sectie als ze aanwezig is, maar behandelt haar expliciet als "
                          "'nog niet gestandaardiseerd' als ze ontbreekt (zie zone 2 in de UI).",
    "Bron": "Notion-pagina 'GroenBuro - Bedrijfscontext' (intern gedeeld met het team)",
    "Laatst_bijgewerkt": None,  # per-bundel ingevuld
    "Placeholders_ingevuld": ["BEDRIJFSNAAM", "DOELGROEP", "SEGMENT_1", "SEGMENT_2", "SEGMENT_3", "POSITIONERING"],
    "Placeholders_open": [],
    "Projectkennis_kopie_laatst_bijgewerkt": None,
}


def build_domain_files():
    """(domein-key -> lijst met rijen) - de canonieke inhoud, ongeacht formaat."""
    return {
        "organisaties": ORGANISATIES,
        "contactpersonen": CONTACTPERSONEN,
        "interacties": INTERACTIES,
        "sales_funnel": SALES_FUNNEL,
        "acties": ACTIES,
        "lessen_inzichten": LESSEN_INZICHTEN,
        "dagverslagen": DAGVERSLAGEN,
        "productbacklog": PRODUCTBACKLOG,
        "klantsucces": KLANTSUCCES,
        "projecten": PROJECTEN,
        "offertes": OFFERTES,
        "content_kalender": CONTENT_KALENDER,
        "delivery_rugzak": DELIVERY_RUGZAK,
        "tijdregistratie": TIJDREGISTRATIE,
        "product_catalogus": PRODUCT_CATALOGUS,
    }


def load_registry_domains():
    schema_path = ROOT / "schema" / "schema.generated.js"
    text = schema_path.read_text(encoding="utf-8")
    json_start = text.index("{")
    payload = json.loads(text[json_start: text.rindex("}") + 1])
    return payload["datadomeinen"]


DOMAIN_NAAM = {}  # key -> registry "naam" (voor sheetnamen/labels)


def kebab(key):
    return key.replace("_", "-")


# ─────────────────────────────────────────────────────────────────────────
# Route 1: Excel-werkboek
#   Gaten: geen 'strategy'/'backoffice' sheets (product_catalogus,
#   tijdregistratie) - dus productbacklog/segmenten wél maar die twee niet;
#   delivery_rugzak leeg (0 rijen); geen bedrijfscontext-tabblad; hele
#   bestand achteraf op 45 dagen oud gezet (drempel is 30) om
#   bundle-brede veroudering te testen; agent "dealmaker" komt nergens in
#   Acties/Lessen voor -> geen spoor.
# ─────────────────────────────────────────────────────────────────────────
def build_excel(domains, registry_domains, out_path):
    wb = Workbook()
    first = True
    included_keys = [k for k in domains.keys() if k not in ("tijdregistratie", "product_catalogus")]
    sheet_meta = []
    for key in included_keys:
        naam = registry_domains[key]["naam"]
        velden = [v["naam"] for v in registry_domains[key]["velden"]]
        rows = domains[key]
        if key == "delivery_rugzak":
            rows = []  # bewust leeg domein
        if first:
            ws = wb.active
            ws.title = naam
            first = False
        else:
            ws = wb.create_sheet(naam)
        ws.append(velden)
        for row in rows:
            ws.append([_cellval(row.get(v, "")) for v in velden])
        sheet_meta.append((naam, velden))

    schema_ws = wb.create_sheet("_schema")
    schema_ws.append(["Tabblad", "Kolommen", "Beschrijving"])
    for naam, velden in sheet_meta:
        schema_ws.append([naam, " · ".join(velden), f"{naam} - testdata, zie testdata/README.md"])

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    # Bestand bewust "verouderen": 45 dagen terug, ouder dan de standaard
    # veroudering-drempel van 30 dagen in het dashboard.
    oud = time.time() - 45 * 86400
    os.utime(out_path, (oud, oud))


def _cellval(v):
    if isinstance(v, list):
        return ", ".join(v)
    if isinstance(v, bool):
        return "Ja" if v else "Nee"
    return v


# ─────────────────────────────────────────────────────────────────────────
# Route 2: data/*.json
#   Vorm per bestand: {"_schema": "...", "items": [...]}  (zoals de
#   bron-intake-fase in core/base/orchestrator/prompt.md beschrijft).
#   Gaten: klantsucces-bestand ontbreekt volledig (delivery-module wel deels
#   aanwezig via projecten/delivery_rugzak, maar klantsucces niet gesynchroniseerd
#   naar deze route); productbacklog.json is op zichzelf verouderd (60 dagen)
#   terwijl de rest van de bundel vers is; bedrijfscontext.json aanwezig en
#   vers (groen zone 2 pad); agent "delivery-architect" heeft geen enkele
#   Acties/Lessen-vermelding -> geen spoor.
# ─────────────────────────────────────────────────────────────────────────
def build_json_bundle(domains, registry_domains, out_dir):
    out_dir.mkdir(parents=True, exist_ok=True)
    skip = {"klantsucces"}
    for key, rows in domains.items():
        if key in skip:
            continue
        naam = registry_domains[key]["naam"]
        schema_desc = naam + " - " + "; ".join(v["naam"] for v in registry_domains[key]["velden"]) + "."
        fname = kebab(key) + ".json"
        payload = {"_schema": schema_desc, "items": rows}
        fpath = out_dir / fname
        fpath.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")

    # bedrijfscontext.json: compleet en vers
    ctx = dict(BEDRIJFSCONTEXT_COMPLEET)
    ctx["Laatst_bijgewerkt"] = d(-4)
    ctx["Projectkennis_kopie_laatst_bijgewerkt"] = d(-4)
    (out_dir / "bedrijfscontext.json").write_text(json.dumps(ctx, indent=2, ensure_ascii=False), encoding="utf-8")

    now = time.time()
    for f in out_dir.glob("*.json"):
        os.utime(f, (now, now))

    # productbacklog.json op zichzelf verouderen (60 dagen oud)
    oud = time.time() - 60 * 86400
    pb = out_dir / "productbacklog.json"
    if pb.exists():
        os.utime(pb, (oud, oud))

# ─────────────────────────────────────────────────────────────────────────
# Route 3: Notion-export
#   Er bestaat nog geen canoniek exportmechanisme (de Coördinator-kant van
#   deze route is nog niet gebouwd) - dit is dus een eigen, expliciet als
#   zodanig gelabeld ontwerp: één JSON-bestand per domein, met dezelfde
#   "_schema" + "items"-vorm als de lokale route, aangevuld met twee
#   Notion-specifieke velden per bestand: "_geexporteerd_op" (wanneer de
#   Coördinator deze export maakte) en "_database_id" (fictief, ter
#   illustratie). Zie testdata/README.md en README.md ("wat ik bewust
#   anders heb gedaan") voor de onderbouwing.
#   Gaten: lessen_inzichten leeg (0 items) -> test "geen lessen = bevinding";
#   bedrijfscontext aanwezig maar verouderd (200 dagen) én met 2 open
#   placeholders -> rood/oranje zone 2-pad; agent "content-strateeg" heeft
#   geen enkele Acties/Lessen-vermelding -> geen spoor.
# ─────────────────────────────────────────────────────────────────────────
def build_notion_export(domains, registry_domains, out_dir):
    out_dir.mkdir(parents=True, exist_ok=True)
    export_moment = d(-1)  # de Coördinator exporteerde gisteren
    for key, rows in domains.items():
        naam = registry_domains[key]["naam"]
        schema_desc = naam + " - " + "; ".join(v["naam"] for v in registry_domains[key]["velden"]) + "."
        if key == "lessen_inzichten":
            rows = []  # bewust leeg domein
        payload = {
            "_schema": schema_desc,
            "_geexporteerd_op": export_moment,
            "_database_id": f"notion-db-fictief-{kebab(key)}",
            "items": rows,
        }
        fname = kebab(key) + ".json"
        (out_dir / fname).write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")

    ctx = dict(BEDRIJFSCONTEXT_COMPLEET)
    ctx["Laatst_bijgewerkt"] = d(-200)
    ctx["Placeholders_ingevuld"] = ["BEDRIJFSNAAM", "DOELGROEP", "SEGMENT_1"]
    ctx["Placeholders_open"] = ["SEGMENT_2", "SEGMENT_3", "POSITIONERING"]
    ctx["Projectkennis_kopie_laatst_bijgewerkt"] = d(-250)  # kopie is OUDER dan de bron -> precies het S17-controlegeval
    ctx["_geexporteerd_op"] = export_moment
    (out_dir / "bedrijfscontext.json").write_text(json.dumps(ctx, indent=2, ensure_ascii=False), encoding="utf-8")

    now = time.time()
    for f in out_dir.glob("*.json"):
        os.utime(f, (now, now))


def main():
    registry_domains = load_registry_domains()
    domains = build_domain_files()

    build_excel(domains, registry_domains, TESTDATA / "agentic-team.xlsx")
    build_json_bundle(domains, registry_domains, TESTDATA / "data")
    build_notion_export(domains, registry_domains, TESTDATA / "notion-export")

    print("Testdata gegenereerd:")
    print(" -", TESTDATA / "agentic-team.xlsx")
    print(" -", TESTDATA / "data")
    print(" -", TESTDATA / "notion-export")


if __name__ == "__main__":
    main()
